from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
from datetime import datetime


def export_clients_excel(clients):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Clientes'
    headers = ['ID', 'Nombre', 'Email', 'Teléfono', 'Edad', 'Peso', 'Altura', 'Meta', 'Estado', 'Registro']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for c in clients:
        ws.append([c.id, f'{c.name} {c.last_name or ""}'.strip(), c.email, c.phone or '', c.age or '', c.weight or '', c.height or '', c.goal or '', c.membership_status or '', c.registration_date.strftime('%d/%m/%Y') if c.registration_date else ''])
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_payments_excel(payments):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pagos'
    headers = ['ID', 'Cliente', 'Monto', 'Método', 'Fecha', 'Descripción']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for p in payments:
        client_name = f'{p.client.name} {p.client.last_name or ""}'.strip() if p.client else '—'
        ws.append([p.id, client_name, p.amount, p.method or '', p.date.strftime('%d/%m/%Y %H:%M') if p.date else '', p.description or ''])
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_checkins_excel(checkins):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Entradas'
    headers = ['ID', 'Cliente', 'Fecha y hora']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for ch in checkins:
        client_name = f'{ch.client.name} {ch.client.last_name or ""}'.strip() if ch.client else '—'
        ws.append([ch.id, client_name, ch.timestamp.strftime('%d/%m/%Y %H:%M') if ch.timestamp else ''])
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
